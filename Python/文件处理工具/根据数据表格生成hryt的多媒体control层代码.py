# -*- coding: utf-8 -*-
import openpyxl
import re

#生成到.cpp文件 委托生成
def generate_text(table):
    lines = []
    for item in table:
        for key, value in item.items():
         line = '    {{ TEXT("{}"), FMessageResponseDelegate::CreateRaw({},&UsbSubCtl::{})}},'.format(key, ' this',value, )
         #line = '   {{ TEXT("{}"), std::bind(&AudioSubCtl::{},{})}},'.format(key, value, ' this')
         lines.append(line) 
    result2 = "\n".join(lines)
    return result2

#生成.h文件
def produce_text(htable):
    hlines = []
    hext = []
    for item in htable:
        for key, value in item.items():
         line = '    DECLARE_DELEGATE_OneParam(FOn{}Event, TSharedPtr<{}t>)'.format(value, key, ' this')
         next = '    FOn{}Event On{};\n'.format(value,value, 'this')
         second = '    void {}();\n'.format(value)
         hlines.append(line)
         hlines.append(next)
         hlines.append(second)
         result2 = "\n".join(hlines)  
    return result2

#生成.cpp文件  函数生成
def producecpp_text(table):
    lines = []
    for item in table:
        for key, value in item.items():
         line = 'void UsbSubCtl::{}() \n{{'.format(value)
         #line = 'void AudioSubCtl::{}() \n{{'.format(value)
         first = '      const FMediaData* Data = static_cast<FMediaData*>(MediaData);\n      if (Data->{}t_Data)\n      {{'.format(key)
         second = '        On{}.ExecuteIfBound(Data->{}t_Data);\n      }}\n}}\n'.format(value, key, ' this')

         lines.append(line)
         lines.append(first)
         lines.append(second)
        
    result2 = "\n".join(lines)
    return result2

#把reques替换为response，/替换为_,头部增加F，_后的第一个字母为大写.  split(),可以方便地将一个字符串按照指定的分隔符拆分成多个部分 
#extend() 函数会将新列表中的所有元素追加到原有的列表中。
def convertString_requestToResponese_replaceUnderScroce_addF_dxChange(input_string):
    parts = input_string.split("/")
    if len(parts) > 1 and parts[1] == "request":
        parts[1] = "Response"
    converted_parts = ["F"]
    for part in parts:
        if part != "":
            converted_parts.extend(part.split("_"))
    for i in range(1, len(converted_parts)):
        converted_parts[i] = converted_parts[i].capitalize()
    output_string = "_".join(converted_parts)
    return output_string

def convert_request_to_response(input_request):
   new_string = input_request
   new_string = input_request.replace("request","response")
   return new_string

def convert2_string(in_string):
    split_result = in_string.split("controller_", 1)
    if len(split_result) > 1:
     extracted_string = split_result[1]
     return(extracted_string)
    else:
     print("Ximalayamediacontroller_未找到目标字符串")
     return 0

#截取字符串第四个/后的内容
def extract_catchXXXXLater(string):
    split_parts = string.split('/')
    if len(split_parts) >= 4:
        return split_parts[4]
    else:
        return ""

#将首字母大写，并将_后的第一个字母大写
def format_string(string):
    words = string.split("_")  # 将字符串按下划线分割成单词列表
    formatted_words = []
    for word in words:
        formatted_word = word.capitalize()
        formatted_words.append(formatted_word)
    formatted_string = "_".join(formatted_words)  # 将单词列表拼接成字符串
    return formatted_string


#读取cpp文件，并生成代码
def replaces_file_string(path,startCode,endCode,replacement):
   with open(path, "r",encoding='utf-8') as file:
        original_content = file.read()

        pattern = re.compile(rf"({re.escape(startCode)})(.*?)(?={re.escape(endCode)})", re.DOTALL)
        result = re.sub(pattern, rf"\1{replacement}", original_content)
   
        file.close()

        with open(path, "w",encoding='utf-8') as filew:
            filew.write(result)
            filew.close()

# 打开Excel文件
workbook = openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\ZXTools\pythonProject\USB.xlsx')

# 根据Sheet名称选择需要读取的Sheet
sheet = workbook['Sheet1']

# 选择要读取的列索引（例如，读取第1列，即A列的数据）
column_index = 1

# 获取列的数据
column_data = [cell.value for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=column_index, max_col=column_index) for cell in row]

#输出列的数据 输出的table里有替换request为response，更改格式后的接口名
Newtable = []
for data in column_data :
    if data is not None:
         converted_data = convertString_requestToResponese_replaceUnderScroce_addF_dxChange(data)
         dataformat = extract_catchXXXXLater(data)
         Newtable.append({
           converted_data:format_string(dataformat) 
           })
hVoidText = produce_text(Newtable)

cppVoidText = producecpp_text(Newtable)

mapTable = []
for data in column_data:
    if data is not None:
        dataformat = extract_catchXXXXLater(data)
        mapTable.append({
            convert_request_to_response(data):format_string(dataformat) 
        })
cppWTVoidText = generate_text(mapTable)

# OptionCpp(r"C:\Users\Administrator\Desktop\pythonCode\cppText.txt",'#xmlyragma endregio','#pragma endregion' , cppTextOutText)

#OptionCpp(r"D:\hryt\HMI\UE4\Source\NIC\UIControl\Media\MediaSubCtl\QQMusicSubCtl.cpp",'#pragma region "根据表格自动生成 QQMusicSubCtl.cpp 函数"','xxx' ,cppVoidText,'#pragma region "根据表格自动生成 QQMusicSubCtl.cpp 委托绑定"','#pragma endregion',cppWTVoidText)

#OptionCpp(r"D:\hryt\HMI\UE4\Source\NIC\UIControl\Media\MediaSubCtl\QQMusicSubCtl.cpp",'#pragma region "根据表格自动生成 QQMusicSubCtl.cpp 函数"','kkf' ,cppVoidText,'#pragma region "根据表格自动生成 QQMusicSubCtl.cpp 委托绑定"','#pragma endregion',cppWTVoidText)
#OptionCpp(r"D:\hryt\HMI\UE4\Source\NIC\UIControl\Media\MediaSubCtl\QQMusicSubCtl.h",'#pragma region "根据表格自动生成 QQMusicSubCtl.h 函数"','#pragma endregion' ,hVoidText,'#pragma region "根据表格自动生成 QQMusicSubCtl.cpp 委托绑定"','#pragma endregion',cppVoidText)


#OptionCpp(r"D:\hryt\HMI\UE4\Source\NIC\UIControl\Media\MediaSubCtl\AudioSubCtl.cpp",'#pragma region "根据表格自动生成 AudioSubCtl.cpp 函数"','kkf' ,cppVoidText,'#pragma region "根据表格自动生成 AudioSubCtl.cpp 委托绑定"','#pragma endregion',cppWTVoidText)
#OptionCpp(r"D:\hryt\HMI\UE4\Source\NIC\UIControl\Media\MediaSubCtl\AudioSubCtl.h",'#pragma region "根据表格自动生成 AudioSubCtl.h 函数"','#pragma endregion' ,hVoidText,'#pragma region "根据表格自动生成 AudioSubCtl.cpp 委托绑定"','#pragma endregion',cppWTVoidText)

# OptionCpp(r"D:\hryt\NIC_UE4_Engine\HMI\UE4\Source\NIC\UIControl\Media\MediaSubCtl\UsbSubCtl.h",'#pragma region "根据表格自动生成 UsbSubCtl.h 函数"','#pragma endregion' ,hVoidText,'#pragma region "根据表格自动生成 UsbSubCtl.cpp 委托绑定"','#pragma endregion',cppVoidText)
# OptionCpp(r"D:\hryt\NIC_UE4_Engine\HMI\UE4\Source\NIC\UIControl\Media\MediaSubCtl\UsbSubCtl.cpp",'#pragma region "根据表格自动生成 UsbSubCtl.cpp 函数"','#pragma endregion' ,cppVoidText,'#pragma region "根据表格自动生成 UsbSubCtl.cpp 委托绑定"','#pragma endregion',cppWTVoidText)


dir = "D:\\hryt\\NIC_UE4_Engine\\HMI\\UE4\\Source\\NIC\\UIControl\\Media\\MediaSubCtl\\"
start_str = '\n#pragma region "{}"\n' 
end_str = '\n#pragma endregion'

replaces_file_string(dir+r"UsbSubCtl.h",start_str.format("根据表格自动生成 UsbSubCtl.h 函数") ,end_str ,hVoidText)

replaces_file_string(dir+r"UsbSubCtl.cpp",start_str.format("根据表格自动生成 UsbSubCtl.cpp 委托绑定"),end_str ,cppWTVoidText)

replaces_file_string(dir+r"UsbSubCtl.cpp",start_str.format("根据表格自动生成 UsbSubCtl.cpp 函数"),end_str ,cppVoidText)
